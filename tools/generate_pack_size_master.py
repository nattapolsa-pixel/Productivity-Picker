"""Generate the browser-side Pack Size lookup from WMS master exports.

Business rule:
- Use QTY, INNERPACK and CASECNT as pickable package sizes.
- Exclude PALLET from normal Pick Productivity.
- Keep unique positive sizes, sorted largest to smallest.
- Add SKU aliases from Item list when SKU and PACKKEY differ.
"""

from __future__ import annotations

import json
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def positive_number(value: object) -> int | float | None:
    try:
        parsed = Decimal(text(value))
    except (InvalidOperation, ValueError):
        return None
    if parsed <= 0:
        return None
    if parsed == parsed.to_integral_value():
        return int(parsed)
    return float(parsed)


def header_indexes(sheet) -> dict[str, int]:
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return {text(value): index for index, value in enumerate(headers) if text(value)}


def load_pack_sizes(path: Path) -> dict[str, list[int | float]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Data"]
    indexes = header_indexes(sheet)
    required = ("PACKKEY", "QTY", "INNERPACK", "CASECNT")
    missing = [name for name in required if name not in indexes]
    if missing:
        raise ValueError(f"Pack Size.xlsx missing columns: {', '.join(missing)}")

    master: dict[str, list[int | float]] = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        pack_key = text(row[indexes["PACKKEY"]])
        if not pack_key:
            continue
        sizes = {
            value
            for field in ("QTY", "INNERPACK", "CASECNT")
            if (value := positive_number(row[indexes[field]])) is not None
        }
        if not sizes:
            continue
        master[pack_key] = sorted(sizes, reverse=True)
    return master


def add_sku_aliases(
    master: dict[str, list[int | float]],
    item_list_path: Path,
) -> tuple[int, int]:
    workbook = load_workbook(item_list_path, read_only=True, data_only=True)
    sheet = workbook["Data"]
    indexes = header_indexes(sheet)
    required = ("SKU", "PACKKEY")
    missing = [name for name in required if name not in indexes]
    if missing:
        raise ValueError(f"Item list.xlsx missing columns: {', '.join(missing)}")

    aliases = 0
    overrides = 0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        sku = text(row[indexes["SKU"]])
        pack_key = text(row[indexes["PACKKEY"]])
        if not sku or not pack_key or sku == pack_key:
            continue
        if pack_key not in master:
            raise ValueError(f"PACKKEY {pack_key} for SKU {sku} is missing from Pack Size.xlsx")
        if sku in master and master[sku] != master[pack_key]:
            overrides += 1
        master[sku] = master[pack_key]
        aliases += 1
    return aliases, overrides


def render_javascript(
    master: dict[str, list[int | float]],
    aliases: int,
    overrides: int,
) -> str:
    rows = [
        f"{json.dumps(key, ensure_ascii=False)}:{json.dumps(master[key], separators=(',', ':'))}"
        for key in sorted(master)
    ]
    return "\n".join(
        [
            "/* Auto-generated from Pack Size.xlsx + Item list.xlsx.",
            "   Normal pick units exclude PALLET and use the largest exact pack divisor. */",
            "globalThis.PACK_SIZE_META = Object.freeze({",
            "  source: 'Pack Size.xlsx + Item list.xlsx',",
            f"  skuCount: {len(master)},",
            f"  skuAliases: {aliases},",
            f"  skuOverrides: {overrides},",
            "  palletExcluded: true",
            "});",
            "globalThis.PACK_SIZE_MASTER = Object.freeze({",
            ",\n".join(rows),
            "});",
            "",
        ]
    )


def main() -> None:
    root = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else Path.cwd()
    output = root / "pack_size_master.js"
    master = load_pack_sizes(root / "Pack Size.xlsx")
    aliases, overrides = add_sku_aliases(master, root / "Item list.xlsx")
    output.write_text(
        render_javascript(master, aliases, overrides),
        encoding="utf-8",
        newline="\n",
    )
    print(
        json.dumps(
            {
                "output": str(output),
                "sku_count": len(master),
                "sku_aliases": aliases,
                "sku_overrides": overrides,
                "target_250001008": master.get("250001008"),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
